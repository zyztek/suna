from __future__ import annotations

import csv
import io
import json
from dataclasses import dataclass
from io import BytesIO
from statistics import mean
from typing import Any, Dict, List, Optional, Tuple

import chardet
from agentpress.tool import ToolResult, openapi_schema, usage_example
from sandbox.tool_base import SandboxToolsBase
from utils.logger import logger

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.chart import BarChart, LineChart, PieChart, Reference, Series, ScatterChart
    from openpyxl.formatting.rule import ColorScaleRule
except Exception:
    openpyxl = None


@dataclass
class SheetData:
    headers: List[str]
    rows: List[List[Any]]


class SandboxSheetsTool(SandboxToolsBase):
    def __init__(self, project_id: str, thread_manager):
        super().__init__(project_id, thread_manager)

    async def _file_exists(self, full_path: str) -> bool:
        try:
            await self.sandbox.fs.get_file_info(full_path)
            return True
        except Exception:
            return False

    async def _download_bytes(self, full_path: str) -> bytes:
        return await self.sandbox.fs.download_file(full_path)

    async def _upload_bytes(self, full_path: str, data: bytes, permissions: str = "644") -> None:
        await self.sandbox.fs.upload_file(data, full_path)
        await self.sandbox.fs.set_file_permissions(full_path, permissions)

    def _detect_encoding(self, data: bytes) -> str:
        try:
            result = chardet.detect(data)
            return result.get("encoding") or "utf-8"
        except Exception:
            return "utf-8"

    def _read_csv_bytes(self, data: bytes) -> SheetData:
        encoding = self._detect_encoding(data)
        text = data.decode(encoding, errors="replace")
        reader = csv.reader(io.StringIO(text))
        rows = [list(r) for r in reader]
        if not rows:
            return SheetData(headers=[], rows=[])
        headers = [str(h) for h in rows[0]]
        data_rows = rows[1:] if len(rows) > 1 else []
        return SheetData(headers=headers, rows=data_rows)

    def _write_csv_bytes(self, sheet: SheetData) -> bytes:
        buf = io.StringIO()
        writer = csv.writer(buf)
        if sheet.headers:
            writer.writerow(sheet.headers)
        for r in sheet.rows:
            writer.writerow(["" if v is None else v for v in r])
        return buf.getvalue().encode("utf-8")

    def _read_xlsx_bytes(self, data: bytes, sheet_name: Optional[str]) -> SheetData:
        if not openpyxl:
            raise RuntimeError("openpyxl not available; cannot read XLSX")
        wb = openpyxl.load_workbook(BytesIO(data), data_only=False)
        ws = wb[sheet_name] if sheet_name else wb.active
        rows = [list(row) for row in ws.iter_rows(values_only=True)]
        if not rows:
            return SheetData(headers=[], rows=[])
        headers = ["" if h is None else str(h) for h in rows[0]]
        data_rows = [[c for c in r] for r in rows[1:]] if len(rows) > 1 else []
        return SheetData(headers=headers, rows=data_rows)

    def _write_xlsx_bytes(self, sheet: SheetData, sheet_name: Optional[str]) -> bytes:
        if not openpyxl:
            raise RuntimeError("openpyxl not available; cannot write XLSX")
        wb = Workbook()
        ws = wb.active
        if sheet_name:
            ws.title = sheet_name
        if sheet.headers:
            ws.append(sheet.headers)
        for r in sheet.rows:
            ws.append([v for v in r])
        out = BytesIO()
        wb.save(out)
        return out.getvalue()

    async def _load_sheet(self, file_path: str, sheet_name: Optional[str]) -> Tuple[str, SheetData]:
        file_path = self.clean_path(file_path)
        full_path = f"{self.workspace_path}/{file_path}"
        data = await self._download_bytes(full_path)
        if file_path.lower().endswith(".csv"):
            return full_path, self._read_csv_bytes(data)
        if file_path.lower().endswith(".xlsx"):
            return full_path, self._read_xlsx_bytes(data, sheet_name)
        raise ValueError("Unsupported file extension. Use .csv or .xlsx")

    async def _save_sheet(self, file_path: str, sheet: SheetData, sheet_name: Optional[str]) -> str:
        file_path = self.clean_path(file_path)
        full_path = f"{self.workspace_path}/{file_path}"
        if file_path.lower().endswith(".csv"):
            await self._upload_bytes(full_path, self._write_csv_bytes(sheet))
        elif file_path.lower().endswith(".xlsx"):
            await self._upload_bytes(full_path, self._write_xlsx_bytes(sheet, sheet_name))
            try:
                csv_full = f"{full_path.rsplit('.', 1)[0]}.csv"
                await self._upload_bytes(csv_full, self._write_csv_bytes(sheet))
            except Exception as e:
                logger.warning(f"Failed to write CSV mirror for {full_path}: {e}")
        else:
            raise ValueError("Unsupported file extension. Use .csv or .xlsx")
        return full_path

    def _infer_column_types(self, rows: List[List[Any]], headers: List[str]) -> Dict[str, str]:
        types: Dict[str, str] = {}
        if not headers:
            return types
        col_count = max((len(r) for r in rows), default=0)
        col_count = max(col_count, len(headers))
        for i in range(col_count):
            col_values = [r[i] for r in rows if len(r) > i]
            detected = "string"
            numeric_count = 0
            date_like = 0
            for v in col_values:
                if isinstance(v, (int, float)):
                    numeric_count += 1
                elif isinstance(v, str):
                    v_strip = v.strip()
                    try:
                        float(v_strip)
                        numeric_count += 1
                        continue
                    except Exception:
                        pass
                    if any(sep in v_strip for sep in ("-", "/")) and any(ch.isdigit() for ch in v_strip):
                        date_like += 1
            if numeric_count >= max(1, len(col_values) // 2):
                detected = "number"
            elif date_like >= max(1, len(col_values) // 2):
                detected = "date"
            types[headers[i] if i < len(headers) else f"col_{i+1}"] = detected
        return types

    def _to_index_map(self, headers: List[str]) -> Dict[str, int]:
        return {h: i for i, h in enumerate(headers)}

    @openapi_schema({
        "type": "function",
        "function": {
            "name": "update_sheet",
            "description": "Modify existing cells, rows, or columns (insert/delete/update).",
            "parameters": {
                "type": "object",
                "properties": {
                    "file_path": {"type": "string"},
                    "sheet_name": {"type": "string", "nullable": True},
                    "operations": {
                        "type": "array",
                        "items": {
                            "type": "object",
                            "properties": {
                                "type": {"type": "string", "enum": [
                                    "update_cell", "update_row", "insert_row", "delete_row", "insert_column", "delete_column"
                                ]},
                                "row_index": {"type": "integer"},
                                "column": {"type": "string"},
                                "column_index": {"type": "integer"},
                                "values": {"type": "array", "items": {"type": "string"}},
                                "value": {"type": "string"}
                            },
                            "required": ["type"]
                        }
                    },
                    "save_as": {"type": "string", "description": "Optional path to save result (.csv or .xlsx)"}
                },
                "required": ["file_path", "operations"]
            }
        }
    })
    @usage_example('''
        <function_calls>
        <invoke name="update_sheet">
          <parameter name="file_path">employee_details.xlsx</parameter>
          <parameter name="sheet_name">Employee Database</parameter>
          <parameter name="operations">[
            {"type":"insert_row","row_index":6,"values":["","","","","","","","","",""]},
            {"type":"insert_row","row_index":7,"values":["SUMMARY CALCULATIONS","","","","","","","","",""]},
            {"type":"insert_row","row_index":8,"values":["Total Employees:","=COUNTA(A2:A100)","","","","","","","",""]}
          ]</parameter>
        </invoke>
        </function_calls>
    ''')
    async def update_sheet(self, file_path: str, operations: List[Dict[str, Any]], sheet_name: Optional[str] = None, save_as: Optional[str] = None) -> ToolResult:
        try:
            await self._ensure_sandbox()
            rel = self.clean_path(file_path)
            full_path = f"{self.workspace_path}/{rel}"

            if rel.lower().endswith(".xlsx"):
                if not openpyxl:
                    return self.fail_response("openpyxl not available to update .xlsx")

                data = await self._download_bytes(full_path)
                wb = openpyxl.load_workbook(BytesIO(data))
                ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active

                header_map: Dict[str, int] = {}
                max_col = ws.max_column or 0
                if ws.max_row >= 1:
                    for c in range(1, max_col + 1):
                        hv = ws.cell(row=1, column=c).value
                        if hv is not None:
                            header_map[str(hv)] = c

                def resolve_col_index(op: Dict[str, Any]) -> Optional[int]:
                    if op.get("column_index"):
                        try:
                            return max(1, int(op["column_index"]))
                        except Exception:
                            return None
                    name = op.get("column")
                    if name and name in header_map:
                        return header_map[name]
                    return None

                for op in operations:
                    t = op.get("type")
                    if t == "update_cell":
                        r = int(op.get("row_index", 0))
                        c = resolve_col_index(op)
                        if r <= 0 or c is None:
                            return self.fail_response("update_cell requires row_index>=1 and column/column_index")
                        val = op.get("value")
                        ws.cell(row=r, column=c).value = val
                        if r == 1:
                            header_map[str(val)] = c
                    elif t == "update_row":
                        r = int(op.get("row_index", 0))
                        if r <= 1:
                            return self.fail_response("update_row requires row_index>=2 (row 1 is header)")
                        vals = op.get("values", [])
                        for idx, v in enumerate(vals, start=1):
                            ws.cell(row=r, column=idx).value = v
                    elif t == "insert_row":
                        r = int(op.get("row_index", 0))
                        if r < 1:
                            r = 1
                        ws.insert_rows(r)
                        vals = op.get("values", [])
                        for idx, v in enumerate(vals, start=1):
                            ws.cell(row=r, column=idx).value = v
                        if r == 1:
                            header_map.clear()
                            max_col = ws.max_column or 0
                            for c in range(1, max_col + 1):
                                hv = ws.cell(row=1, column=c).value
                                if hv is not None:
                                    header_map[str(hv)] = c
                    elif t == "delete_row":
                        r = int(op.get("row_index", 0))
                        if r < 1:
                            continue
                        ws.delete_rows(r)
                        if r == 1:
                            header_map.clear()
                            max_col = ws.max_column or 0
                            if ws.max_row >= 1:
                                for c in range(1, max_col + 1):
                                    hv = ws.cell(row=1, column=c).value
                                    if hv is not None:
                                        header_map[str(hv)] = c
                    elif t == "insert_column":
                        c = resolve_col_index(op)
                        if c is None:
                            c = (ws.max_column or 0) + 1
                        ws.insert_cols(c)
                        new_header = op.get("column")
                        if new_header:
                            ws.cell(row=1, column=c).value = new_header
                            header_map[str(new_header)] = c
                    elif t == "delete_column":
                        c = resolve_col_index(op)
                        if c is None:
                            continue
                        ws.delete_cols(c)
                        header_map.clear()
                        max_col = ws.max_column or 0
                        if ws.max_row >= 1:
                            for ci in range(1, max_col + 1):
                                hv = ws.cell(row=1, column=ci).value
                                if hv is not None:
                                    header_map[str(hv)] = ci
                    else:
                        return self.fail_response(f"Unsupported operation type: {t}")

                out = BytesIO()
                wb.save(out)
                await self._upload_bytes(full_path if not save_as else f"{self.workspace_path}/{self.clean_path(save_as)}", out.getvalue())
                try:
                    csv_full = f"{(full_path if not save_as else f'{self.workspace_path}/{self.clean_path(save_as)}').rsplit('.', 1)[0]}.csv"
                    from csv import writer as csv_writer
                    csv_buf = io.StringIO()
                    w = csv_writer(csv_buf)
                    for r in ws.iter_rows(values_only=True):
                        w.writerow(list(r))
                    await self._upload_bytes(csv_full, csv_buf.getvalue().encode('utf-8'))
                except Exception:
                    pass

                saved_path = (save_as or file_path)
                return self.success_response({"updated": f"{self.workspace_path}/{self.clean_path(saved_path)}", "headers": [ws.cell(row=1, column=c).value for c in range(1, (ws.max_column or 0)+1)], "row_count": ws.max_row})

            full_path, sheet = await self._load_sheet(file_path, sheet_name)

            headers = sheet.headers[:] or []
            index_map = self._to_index_map(headers) if headers else {}

            def resolve_col(op: Dict[str, Any]) -> Optional[int]:
                if op.get("column_index"):
                    return max(1, int(op["column_index"])) - 1
                name = op.get("column")
                if name and name in index_map:
                    return index_map[name]
                return None

            for op in operations:
                t = op.get("type")
                if t == "update_cell":
                    r_idx = int(op.get("row_index", 0)) - 1
                    c_idx = resolve_col(op)
                    if r_idx < 0 or c_idx is None:
                        return self.fail_response("update_cell requires row_index>=1 and column/column_index")
                    if r_idx == 0:
                        if not headers:
                            return self.fail_response("Cannot update header without headers present.")
                        if c_idx >= len(headers):
                            headers.extend([""] * (c_idx - len(headers) + 1))
                        headers[c_idx] = op.get("value")
                        index_map = self._to_index_map(headers)
                    else:
                        data_idx = r_idx - 1
                        while data_idx >= len(sheet.rows):
                            sheet.rows.append([None] * max(1, len(headers)))
                        row = sheet.rows[data_idx]
                        if c_idx >= len(row):
                            row.extend([None] * (c_idx - len(row) + 1))
                        row[c_idx] = op.get("value")
                elif t == "update_row":
                    r_idx = int(op.get("row_index", 0)) - 1
                    if r_idx <= 0:
                        return self.fail_response("update_row requires row_index>=2 (row 1 is header)")
                    data_idx = r_idx - 1
                    vals = op.get("values", [])
                    while data_idx >= len(sheet.rows):
                        sheet.rows.append([None] * max(1, len(headers)))
                    sheet.rows[data_idx] = vals
                elif t == "insert_row":
                    r_idx = int(op.get("row_index", 0)) - 1
                    if r_idx < 0:
                        r_idx = 0
                    if r_idx == 0:
                        headers = [str(v) for v in op.get("values", [])]
                        index_map = self._to_index_map(headers)
                    else:
                        data_idx = max(0, r_idx - 1)
                        while data_idx > len(sheet.rows):
                            sheet.rows.append([None] * max(1, len(headers)))
                        sheet.rows.insert(data_idx, op.get("values", []))
                elif t == "delete_row":
                    r_idx = int(op.get("row_index", 0)) - 1
                    if r_idx == 0:
                        headers = []
                        index_map = {}
                    else:
                        data_idx = r_idx - 1
                        if 0 <= data_idx < len(sheet.rows):
                            sheet.rows.pop(data_idx)
                elif t == "insert_column":
                    c_idx = resolve_col(op)
                    if c_idx is None:
                        c_idx = len(headers)
                    new_header = op.get("column", f"col_{c_idx+1}")
                    if not headers:
                        headers = [new_header]
                    else:
                        if c_idx > len(headers):
                            headers.extend([""] * (c_idx - len(headers)))
                        headers.insert(c_idx, new_header)
                    for i in range(len(sheet.rows)):
                        row = sheet.rows[i]
                        if c_idx > len(row):
                            row.extend([None] * (c_idx - len(row)))
                        row.insert(c_idx, None)
                    index_map = self._to_index_map(headers)
                elif t == "delete_column":
                    c_idx = resolve_col(op)
                    if c_idx is None or not headers or c_idx >= len(headers):
                        continue
                    headers.pop(c_idx)
                    for row in sheet.rows:
                        if c_idx < len(row):
                            row.pop(c_idx)
                    index_map = self._to_index_map(headers)
                else:
                    return self.fail_response(f"Unsupported operation type: {t}")

            sheet.headers = headers

            target_path = save_as or file_path
            saved_path = await self._save_sheet(target_path, sheet, sheet_name)
            return self.success_response({"updated": saved_path, "row_count": len(sheet.rows), "headers": sheet.headers})
        except Exception as e:
            logger.exception("update_sheet failed")
            return self.fail_response(f"Error updating sheet: {e}")

    @openapi_schema({
        "type": "function",
        "function": {
            "name": "view_sheet",
            "description": "Read headers, types, and sample rows; optional CSV export.",
            "parameters": {
                "type": "object",
                "properties": {
                    "file_path": {"type": "string"},
                    "sheet_name": {"type": "string", "nullable": True},
                    "max_rows": {"type": "integer", "default": 100},
                    "export_csv_path": {"type": "string"}
                },
                "required": ["file_path"]
            }
        }
    })
    @usage_example('''
        <function_calls>
        <invoke name="view_sheet">
          <parameter name="file_path">reports/sales.xlsx</parameter>
          <parameter name="sheet_name">Sheet1</parameter>
          <parameter name="max_rows">50</parameter>
          <parameter name="export_csv_path">reports/sales_preview.csv</parameter>
        </invoke>
        </function_calls>
    ''')
    async def view_sheet(self, file_path: str, sheet_name: Optional[str] = None, max_rows: int = 100, export_csv_path: Optional[str] = None) -> ToolResult:
        try:
            await self._ensure_sandbox()
            full_path, sheet = await self._load_sheet(file_path, sheet_name)
            exported_to = None
            if export_csv_path:
                rel = self.clean_path(export_csv_path)
                if not rel.lower().endswith(".csv"):
                    rel += ".csv"
                export_full = f"{self.workspace_path}/{rel}"
                await self._upload_bytes(export_full, self._write_csv_bytes(sheet))
                exported_to = export_full
            sample_rows = sheet.rows[: max(0, max_rows)]
            return self.success_response({
                "file_path": full_path,
                "headers": sheet.headers,
                "row_count": len(sheet.rows),
                "sample_rows": sample_rows,
                "exported_csv": exported_to
            })
        except Exception as e:
            logger.exception("view_sheet failed")
            return self.fail_response(f"Error viewing sheet: {e}")

    @openapi_schema({
        "type": "function",
        "function": {
            "name": "create_sheet",
            "description": "Create a new CSV/XLSX with optional headers/rows.",
            "parameters": {
                "type": "object",
                "properties": {
                    "file_path": {"type": "string"},
                    "headers": {"type": "array", "items": {"type": "string"}},
                    "rows": {"type": "array", "items": {"type": "array", "items": {"type": "string"}}},
                    "sheet_name": {"type": "string", "nullable": True},
                    "overwrite": {"type": "boolean", "default": False}
                },
                "required": ["file_path"]
            }
        }
    })
    @usage_example('''
        <function_calls>
        <invoke name="create_sheet">
          <parameter name="file_path">data/sample.csv</parameter>
          <parameter name="headers">["region","revenue"]</parameter>
          <parameter name="rows">[["NA",100],["EU",120]]</parameter>
          <parameter name="overwrite">true</parameter>
        </invoke>
        </function_calls>
    ''')
    async def create_sheet(self, file_path: str, headers: Optional[List[str]] = None, rows: Optional[List[List[Any]]] = None, sheet_name: Optional[str] = None, overwrite: bool = False) -> ToolResult:
        try:
            await self._ensure_sandbox()
            rel = self.clean_path(file_path)
            full = f"{self.workspace_path}/{rel}"
            exists = await self._file_exists(full)
            if exists and not overwrite:
                return self.fail_response("File already exists. Set overwrite=true to replace.")
            if rel.lower().endswith(".csv"):
                await self._upload_bytes(full, self._write_csv_bytes(SheetData(headers or [], rows or [])))
            elif rel.lower().endswith(".xlsx"):
                if not openpyxl:
                    return self.fail_response("openpyxl not available to create .xlsx")
                sheet = SheetData(headers or [], rows or [])
                await self._upload_bytes(full, self._write_xlsx_bytes(sheet, sheet_name))
                try:
                    csv_full = f"{full.rsplit('.', 1)[0]}.csv"
                    await self._upload_bytes(csv_full, self._write_csv_bytes(sheet))
                except Exception as e:
                    logger.warning(f"Failed to write CSV mirror for {full}: {e}")
            else:
                return self.fail_response("Unsupported extension. Use .csv or .xlsx")
            return self.success_response({"created": full, "rows": len(rows or []), "headers": headers or []})
        except Exception as e:
            logger.exception("create_sheet failed")
            return self.fail_response(f"Error creating sheet: {e}")

    @openapi_schema({
        "type": "function",
        "function": {
            "name": "analyze_sheet",
            "description": "Simple statistics and optional group_by; can export CSV.",
            "parameters": {
                "type": "object",
                "properties": {
                    "file_path": {"type": "string"},
                    "sheet_name": {"type": "string", "nullable": True},
                    "target_columns": {"type": "array", "items": {"type": "string"}},
                    "group_by": {"type": "string"},
                    "aggregations": {"type": "array", "items": {"type": "string", "enum": ["count", "sum", "avg", "min", "max"]}},
                    "export_csv_path": {"type": "string"}
                },
                "required": ["file_path"]
            }
        }
    })
    @usage_example('''
        <function_calls>
        <invoke name="analyze_sheet">
          <parameter name="file_path">data/sales.csv</parameter>
          <parameter name="target_columns">["revenue"]</parameter>
          <parameter name="group_by">region</parameter>
          <parameter name="export_csv_path">data/sales_summary.csv</parameter>
        </invoke>
        </function_calls>
    ''')
    async def analyze_sheet(self, file_path: str, sheet_name: Optional[str] = None, target_columns: Optional[List[str]] = None, group_by: Optional[str] = None, aggregations: Optional[List[str]] = None, export_csv_path: Optional[str] = None) -> ToolResult:
        try:
            await self._ensure_sandbox()
            full_path, sheet = await self._load_sheet(file_path, sheet_name)
            headers = sheet.headers
            idx_map = self._to_index_map(headers)

            def to_float(v: Any) -> Optional[float]:
                if v is None:
                    return None
                if isinstance(v, (int, float)):
                    return float(v)
                try:
                    return float(str(v).strip())
                except Exception:
                    return None

            numeric_cols = [c for c in (target_columns or headers) if c in idx_map]
            if group_by and group_by in idx_map:
                g_idx = idx_map[group_by]
                groups: Dict[Any, List[List[Any]]] = {}
                for row in sheet.rows:
                    key = row[g_idx] if len(row) > g_idx else None
                    groups.setdefault(key, []).append(row)
                out_headers = [group_by]
                aggs = aggregations or ["count", "sum", "avg", "min", "max"]
                for col in numeric_cols:
                    for agg in aggs:
                        out_headers.append(f"{col}_{agg}")
                summary_rows: List[List[Any]] = []
                for key, rows in groups.items():
                    row_out = [key]
                    for col in numeric_cols:
                        c_idx = idx_map[col]
                        vals = [to_float(r[c_idx]) for r in rows if len(r) > c_idx]
                        vals = [v for v in vals if v is not None]
                        count_v = len(vals)
                        sum_v = sum(vals) if vals else None
                        avg_v = mean(vals) if vals else None
                        min_v = min(vals) if vals else None
                        max_v = max(vals) if vals else None
                        for agg in aggs:
                            row_out.append({
                                "count": count_v,
                                "sum": sum_v,
                                "avg": avg_v,
                                "min": min_v,
                                "max": max_v
                            }[agg])
                    summary_rows.append(row_out)
                result_sheet = SheetData(headers=out_headers, rows=summary_rows)
            else:
                out_headers = ["metric"] + numeric_cols
                rows_out: List[List[Any]] = []
                counts = []
                for col in numeric_cols:
                    c_idx = idx_map[col]
                    vals = [to_float(r[c_idx]) for r in sheet.rows if len(r) > c_idx]
                    vals = [v for v in vals if v is not None]
                    counts.append(len(vals))
                rows_out.append(["count", *counts])
                sums = []
                for col in numeric_cols:
                    c_idx = idx_map[col]
                    vals = [to_float(r[c_idx]) for r in sheet.rows if len(r) > c_idx]
                    vals = [v for v in vals if v is not None]
                    sums.append(sum(vals) if vals else None)
                rows_out.append(["sum", *sums])
                avgs = []
                for col in numeric_cols:
                    c_idx = idx_map[col]
                    vals = [to_float(r[c_idx]) for r in sheet.rows if len(r) > c_idx]
                    vals = [v for v in vals if v is not None]
                    avgs.append(mean(vals) if vals else None)
                rows_out.append(["avg", *avgs])
                mins = []
                for col in numeric_cols:
                    c_idx = idx_map[col]
                    vals = [to_float(r[c_idx]) for r in sheet.rows if len(r) > c_idx]
                    vals = [v for v in vals if v is not None]
                    mins.append(min(vals) if vals else None)
                rows_out.append(["min", *mins])
                maxs = []
                for col in numeric_cols:
                    c_idx = idx_map[col]
                    vals = [to_float(r[c_idx]) for r in sheet.rows if len(r) > c_idx]
                    vals = [v for v in vals if v is not None]
                    maxs.append(max(vals) if vals else None)
                rows_out.append(["max", *maxs])
                result_sheet = SheetData(headers=out_headers, rows=rows_out)

            exported = None
            if export_csv_path:
                rel = self.clean_path(export_csv_path)
                if not rel.lower().endswith(".csv"):
                    rel += ".csv"
                export_full = f"{self.workspace_path}/{rel}"
                await self._upload_bytes(export_full, self._write_csv_bytes(result_sheet))
                exported = export_full

            return self.success_response({
                "analyzed_from": full_path,
                "result_preview": {"headers": result_sheet.headers, "rows": result_sheet.rows[:50]},
                "exported_csv": exported
            })
        except Exception as e:
            logger.exception("analyze_sheet failed")
            return self.fail_response(f"Error analyzing sheet: {e}")

    @openapi_schema({
        "type": "function",
        "function": {
            "name": "visualize_sheet",
            "description": "Generate charts (bar, line, pie, scatter) and save to XLSX. Also optionally export chart data to CSV for Google Sheets.",
            "parameters": {
                "type": "object",
                "properties": {
                    "file_path": {"type": "string"},
                    "sheet_name": {"type": "string", "nullable": True},
                    "chart_type": {"type": "string", "enum": ["bar", "line", "pie", "scatter"], "default": "bar"},
                    "x_column": {"type": "string"},
                    "y_columns": {"type": "array", "items": {"type": "string"}},
                    "save_as": {"type": "string"},
                    "export_csv_path": {"type": "string", "description": "Optional CSV path for the chart dataset (x + y columns)"}
                },
                "required": ["file_path", "x_column", "y_columns"]
            }
        }
    })
    @usage_example('''
        <function_calls>
        <invoke name="visualize_sheet">
          <parameter name="file_path">reports/sales.csv</parameter>
          <parameter name="x_column">region</parameter>
          <parameter name="y_columns">["revenue"]</parameter>
          <parameter name="chart_type">bar</parameter>
          <parameter name="save_as">reports/sales_chart.xlsx</parameter>
          <parameter name="export_csv_path">reports/sales_chart_data.csv</parameter>
        </invoke>
        </function_calls>
    ''')
    async def visualize_sheet(self, file_path: str, x_column: str, y_columns: List[str], chart_type: str = "bar", sheet_name: Optional[str] = None, save_as: Optional[str] = None, export_csv_path: Optional[str] = None) -> ToolResult:
        try:
            await self._ensure_sandbox()
            rel = self.clean_path(file_path)
            full = f"{self.workspace_path}/{rel}"
            _, sheet = await self._load_sheet(file_path, sheet_name)
            headers = sheet.headers
            idx_map = self._to_index_map(headers)
            if x_column not in idx_map:
                return self.fail_response(f"x_column '{x_column}' not found")
            for yc in y_columns:
                if yc not in idx_map:
                    return self.fail_response(f"y_column '{yc}' not found")

            target = save_as or (rel.rsplit(".", 1)[0] + "_chart.xlsx")
            if not target.lower().endswith(".xlsx"):
                target += ".xlsx"
            target_full = f"{self.workspace_path}/{self.clean_path(target)}"
            if not openpyxl:
                return self.fail_response("openpyxl not available to build charts")

            wb = Workbook()
            ws = wb.active
            ws.title = sheet_name or "Data"
            if headers:
                ws.append(headers)
            for r in sheet.rows:
                ws.append(r)

            if chart_type == "bar":
                chart = BarChart()
            elif chart_type == "line":
                chart = LineChart()
            elif chart_type == "pie":
                chart = PieChart()
            else:
                chart = ScatterChart()

            x_col_idx = idx_map[x_column] + 1
            y_col_indices = [idx_map[c] + 1 for c in y_columns]
            min_row = 2
            max_row = len(sheet.rows) + 1
            x_ref = Reference(ws, min_col=x_col_idx, min_row=min_row, max_row=max_row)

            if chart_type == "pie" and len(y_col_indices) == 1:
                data_ref = Reference(ws, min_col=y_col_indices[0], min_row=1, max_row=max_row)
                chart.add_data(data_ref, titles_from_data=True)
                chart.set_categories(x_ref)
            else:
                for yci in y_col_indices:
                    data_ref = Reference(ws, min_col=yci, min_row=min_row - 1, max_row=max_row)
                    series = Series(data_ref, title_from_data=True)
                    series.category = x_ref
                    if isinstance(chart, ScatterChart):
                        series.xvalues = x_ref
                    chart.series.append(series)

            chart_ws = wb.create_sheet(title=f"Chart_{chart_type}")
            chart_ws.add_chart(chart, "A1")
            out = BytesIO()
            wb.save(out)
            await self._upload_bytes(target_full, out.getvalue())

            dataset_headers = [x_column] + y_columns
            dataset_rows: List[List[Any]] = []
            x_idx = idx_map[x_column]
            y_idx_list = [idx_map[y] for y in y_columns]
            for row in sheet.rows:
                if len(row) <= x_idx:
                    continue
                vals = [row[x_idx]]
                ok = True
                for yi in y_idx_list:
                    if len(row) <= yi:
                        ok = False
                        break
                    vals.append(row[yi])
                if ok:
                    dataset_rows.append(vals)

            csv_rel = None
            if export_csv_path:
                csv_rel = self.clean_path(export_csv_path)
                if not csv_rel.lower().endswith(".csv"):
                    csv_rel += ".csv"
            else:
                base = self.clean_path(target).rsplit(".", 1)[0]
                csv_rel = f"{base}_data.csv"
            csv_full = f"{self.workspace_path}/{csv_rel}"
            await self._upload_bytes(csv_full, self._write_csv_bytes(SheetData(headers=dataset_headers, rows=dataset_rows)))

            return self.success_response({
                "source": full,
                "chart_saved": target_full,
                "chart_type": chart_type,
                "chart_data_csv": csv_full
            })
        except Exception as e:
            logger.exception("visualize_sheet failed")
            return self.fail_response(f"Error visualizing sheet: {e}")

    @openapi_schema({
        "type": "function",
        "function": {
            "name": "format_sheet",
            "description": "Style and format sheet cells (bold headers, auto widths, optional conditional formatting). XLSX only.",
            "parameters": {
                "type": "object",
                "properties": {
                    "file_path": {"type": "string"},
                    "sheet_name": {"type": "string", "nullable": True},
                    "bold_headers": {"type": "boolean", "default": True},
                    "auto_width": {"type": "boolean", "default": True},
                    "apply_banding": {"type": "boolean", "default": True},
                    "conditional_format": {"type": "object", "properties": {
                        "column": {"type": "string"},
                        "min_color": {"type": "string", "default": "FFEFEB"},
                        "mid_color": {"type": "string", "default": "FFD7D2"},
                        "max_color": {"type": "string", "default": "FFA39E"}
                    }}
                },
                "required": ["file_path"]
            }
        }
    })
    @usage_example('''
        <function_calls>
        <invoke name="format_sheet">
          <parameter name="file_path">reports/sales_chart.xlsx</parameter>
          <parameter name="sheet_name">Data</parameter>
          <parameter name="bold_headers">true</parameter>
          <parameter name="conditional_format">{"column":"revenue"}</parameter>
        </invoke>
        </function_calls>
    ''')
    async def format_sheet(self, file_path: str, sheet_name: Optional[str] = None, bold_headers: bool = True, auto_width: bool = True, apply_banding: bool = True, conditional_format: Optional[Dict[str, Any]] = None) -> ToolResult:
        try:
            await self._ensure_sandbox()
            rel = self.clean_path(file_path)
            full = f"{self.workspace_path}/{rel}"
            if not rel.lower().endswith(".xlsx"):
                return self.fail_response("format_sheet only supports .xlsx")
            data = await self._download_bytes(full)
            if not openpyxl:
                return self.fail_response("openpyxl not available")
            wb = openpyxl.load_workbook(BytesIO(data))
            ws = wb[sheet_name] if sheet_name else wb.active

            max_col = ws.max_column
            max_row = ws.max_row

            if bold_headers and max_row >= 1:
                for c in range(1, max_col + 1):
                    cell = ws.cell(row=1, column=c)
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(vertical="center")

            if apply_banding and max_row > 2:
                for r in range(2, max_row + 1):
                    if r % 2 == 0:
                        for c in range(1, max_col + 1):
                            ws.cell(row=r, column=c).fill = PatternFill(start_color="FFF9F9", end_color="FFF9F9", fill_type="solid")

            if auto_width:
                for c in range(1, max_col + 1):
                    max_len = 0
                    for r in range(1, max_row + 1):
                        v = ws.cell(row=r, column=c).value
                        max_len = max(max_len, len(str(v)) if v is not None else 0)
                    ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = min(60, max(10, max_len + 2))

            if conditional_format and max_row > 1:
                column_name = conditional_format.get("column")
                if column_name:
                    headers = [ws.cell(row=1, column=c).value for c in range(1, max_col + 1)]
                    if column_name in headers:
                        c_idx = headers.index(column_name) + 1
                        rng = f"{openpyxl.utils.get_column_letter(c_idx)}2:{openpyxl.utils.get_column_letter(c_idx)}{max_row}"
                        ws.conditional_formatting.add(
                            rng,
                            ColorScaleRule(start_type='min', start_color=conditional_format.get("min_color", "FFEFEB"),
                                           mid_type='percentile', mid_value=50, mid_color=conditional_format.get("mid_color", "FFD7D2"),
                                           end_type='max', end_color=conditional_format.get("max_color", "FFA39E"))
                        )

            out = BytesIO()
            wb.save(out)
            await self._upload_bytes(full, out.getvalue())
            return self.success_response({"formatted": full, "sheet": ws.title})
        except Exception as e:
            logger.exception("format_sheet failed")
            return self.fail_response(f"Error formatting sheet: {e}") 